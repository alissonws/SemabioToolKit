#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

import os
# módulo para enviar emails (incluindo emails simples e com anexos)
import smtplib, platform
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

from pathlib import Path
import time


class EmailsSBChecking:
    def __init__(self):
        pass
    def check_plan(self, dir,col_email="",headers = 0):
        """
        Método para checar a viabilidade da planilha e/ou contar o número de e-mails.
        Sem os dois últimos parâmetros, faz o primeiro, com todos os parâmetros preenchidos, faz o segundo.
        :param dir: diretório da planilha;
        :param col_email: coluna da planilha com os e-mails;
        :param headers: considerar a primeira linha como cabeçalho da planilha;
        :return:
        """

        try:
            wb = load_workbook(dir)
            ws = wb.active

            if col_email != "":
                mount = len(ws[col_email]) - headers
                return mount



        except: raise ImportError("Planilha invállida")

    def check_login(self, email, pwd):
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.ehlo()
        server.starttls()
        server.login(email, pwd)
        server.quit()
        return True


    def preview(self, plan_dir, subject, body, col_email, header = 0, col_annex = "", annex_folder = "", annex_ext = "", col_args = {}):
        #carregando planilha
        wb = load_workbook(plan_dir)
        ws = wb.active

        start = 1 + header
        count=0

        #carregando planilha
        toaddr = ws[col_email + str(count + start)].value


        # salvando uma versão do corpo do emails para edição
        subject_new = subject
        body_new = body
        #substituindo tags
        for tuple in col_args:
            cell = tuple[0]
            tag = tuple[1]

            print(cell, " e ", tag)

            if ws[cell + str(start + count)].value == None:
                skip = 1
                break

            subject_new = subject_new.replace(tag, str(ws[cell + str(start + count)].value))
            body_new = body_new.replace(tag, str(ws[cell + str(start + count)].value))

        #anexo

        if col_annex != "":
            filename = Path(str(ws[col_annex + str(header + 1)].value))

            if filename.suffix != annex_ext and annex_ext != "":
                filename = filename.with_suffix(annex_ext)

            path_to_annex = Path(annex_folder) / filename
            try:
                with open(path_to_annex) as file:
                    file.close()
            except FileNotFoundError:
                filename = "ERRO-1"

        #Concatenando resultados
        preview = [toaddr,subject_new,body_new]
        if "filename" in locals():
            preview.append(str(filename))

        return preview


class EmailsSB():
    def __init__(self, queue_d, queue_p, plan_dir, fromaddr, pwd, subject, body, col_email, header = 0, col_annex = "", annex_folder = "", annex_ext = "", col_args = {}):

        #Contando tempo de execução
        start_time = time.time()
        print("DIr: %s"% plan_dir)
        print("fromaddr: %s"% fromaddr)
        print("pwd %s"%pwd)

        #carregando planilha
        wb = load_workbook(plan_dir)
        ws = wb.active

        #criar lista com os emails a enviar
        lim = len(ws[col_email])
        start = 1 + header
        count=0
        print(count)

        # conexão base com o servidor
        # para outros emails, buscar o servidor e porta

        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.ehlo()
        server.starttls()
        server.login(fromaddr, pwd)

        print("Conactado.")


        while count < lim:

            toaddr = ws[col_email + str(count+start)].value

            # corpo de texto
            msg = MIMEMultipart()

            msg["From"] = fromaddr
            msg["To"] = toaddr


            #salvando uma versão do corpo do emails para edição
            subject_new = subject
            body_new = body
            #substituindo tags

            for tuple in col_args:
                cell = tuple[0]
                tag = tuple[1]

                print(cell," e ",tag)

                if ws[cell+str(start+count)].value == None:
                    skip = 1
                    break

                subject_new = subject_new.replace(tag, str(ws[cell + str(start+count)].value))
                body_new = body_new.replace(tag, str(ws[cell + str(start+count)].value))




            #adicionando valores editados ao email

            msg["Subject"] = subject_new
            msg.attach(MIMEText(body_new, "plain"))

            #anexo
            # é necessário converter o anexo para base64 antes de enviar

            if col_annex != "":

                filename = Path(str(ws[col_annex+str(start+count)].value))

                if filename.suffix != annex_ext and annex_ext != "":
                    filename = filename.with_suffix(annex_ext)

                path_to_annex = Path(annex_folder) / filename

                attachment = open(path_to_annex, "rb")

                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", "attachment; filename= %s" % filename)
                msg.attach(part)
                attachment.close()

            #passando informações para a janela de atividade
            data = [toaddr,subject_new,body_new]
            if "filename" in locals():
                data.append(str(filename))

            queue_d.put(data)

            text = msg.as_string()
            print("Enviando email para %s" % toaddr)
            #time.sleep(2)
            server.sendmail(fromaddr, toaddr, text)
            progress = (count/lim)*100
            print("Concluido\nProgresso: "+str(progress)+"%")

            queue_p.put(progress)

            count +=1

        server.quit()
        wb.save(plan_dir)
        print("Terminado")
        print("Tempo de execução para "+str(lim)+" envios: "+str(time.time()-start_time))




if __name__ == "__main__":


    certs_folder = "/home/alisson/certs/"
    plan_dir = "/home/alisson/Área de Trabalho/Controle ministrantes de minicurso.xlsx"

    stop = 0
    wb = load_workbook(plan_dir)
    ws = wb.active
    index = os.listdir(certs_folder)
    ind = []

    for a in index:
        a = a.replace(".pdf","")
        ind.append(a)

    while stop == 0:
        for i in ind:
            if ws["H"+str(i)].value == None:
                nome = ws["D" + str(i)].value
                email = ws["E" + str(i)].value
                fromaddr = "inscricoesdasemabio@gmail.com"
                pwd = "inscricoessemabio"
                toaddr = email

                print("Enviando e-mail para "+ str(nome)+ ": "+ str(email))
                #corpo de texto
                msg = MIMEMultipart()

                msg["From"] = fromaddr
                msg["To"] = toaddr
                msg["Subject"] = "Certificado de ministrante de minicurso na XIX Semana Acadêmica da Biologia"

                body = "Olá, "+str(nome.split(" ")[0])+".\n"+\
                "Segue anexo o seu certificado de ministrante de minicurso na XIX Semana Acadêmica da Biologia."+\
                " Nós, da Comissão Organizadora, agradecemos a sua participação e sua contribuição com o nosso evento."+ \
                " Qualquer erro ou dúvida, nos comunique o mais rápido possível." + \
                "\n\nAtenciosamente,\nAlisson Corrêa,\nComissão de Inscrições e Certificados."

                msg.attach(MIMEText(body, "plain"))


                #é necessário converter o anexo para base64 antes de enviar
                filename = str(i)+".pdf"
                attachment = open(str(str(certs_folder) + str(i) + ".pdf"), "rb")

                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", "attachment; filename= %s" % filename)
                msg.attach(part)

                #conexão base com o servidor
                #para outros emails, buscar o servidor e porta


                server = smtplib.SMTP("smtp.gmail.com", 587)
                server.ehlo()
                server.starttls()
                server.login(fromaddr, pwd)
                text = msg.as_string()
                server.sendmail(fromaddr, toaddr, text)
                server.quit()
                print("Enviado")

                ws["H"+str(i)].value = "X"


                if i == ind[-1]:
                    stop = 1
            else: print("Certificado de "+str(ws["D" + str(i)].value)+"já enviado.")

            if i == ind[-1]:
                stop = 1
    print("Terminado")
    wb.save(plan_dir)